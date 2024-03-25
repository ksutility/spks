
import k_err
from openpyxl import Workbook,load_workbook
'''
wb = Workbook() 
wb = load_workbook('test.xlsx')
wb.save('new_document.xlsm')

#worksheet
ws = wb.active  # grab the active worksheet

#base inf
ws.calculate_dimension()
ws.title

#cell value
print (ws['E3'].value)
ws.cell(row=4, column=2)=5
ws['X4']=ws['E4'].value
d = ws.cell(row=4, column=2, value=10)

#cell other
ws['A1'].number_format='yyyy-mm-dd h:mm:ss'

#end
wb.save(path1)

'''
def wb(wb_path):
    return( load_workbook( wb_path ))
def save_wb(wb,wb_path=''):
    '''
        02/04/20
        goal
            ذخیره فایل اکسل به صورت هوشمند همراه با
                ساخت فلدر فایل مورد نظر در صورت عدم وجود
                مدیریت خطاهایی که باعث عدم ذخیره کامل فایل می شود مثل باز بودن فایل یا فقط خواندنی بودن
    '''
    import k_file,os
    import tk_ui as ui
    if wb_path=='':wb_path=k_file.name_append(wb.path,"-new") 
    
    ff=k_file.file_name_split(wb_path)
    print(str(k_file.dir_make(ff['path'])))
    
    #if not os.access(wb_path, os.W_OK):
    
    while True:
        try:
            wb.save(wb_path)
            print ('ok save =>'+wb_path)
            return True
        except:
            print ('error save =>'+wb_path)
            wp1=k_file.name_append(wb_path,"$")
            x=ui.ask(f'this excel file can not save:\n- {wb_path}<hr>you can Rename it in this path:\n- {wp1}',['retry','rename','cancel'])
            if x=='rename':
                wb_path=wp1
            elif x=='retry':
                pass
            else:
                return False
def open_sheet(wb_path,sheet_name):
    '''
        02/04/20
        goal
            باز کردن یک شیت اکسل به صورت هوشمند همراه با
                ساخت شیت در صورت عدم وجود
                ساخت فایل در صورت عدم وجود
    '''
    import os
    if os.path.exists(wb_path):
        wb = load_workbook( wb_path )
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name) 
        file_is_new=False    
    else:
        wb = Workbook()
        ws= wb.active
        ws.title = sheet_name
        file_is_new=True
    return  (wb,ws,file_is_new)
def get_col_names(x_ws):
    return [str(x.value).lower() for x in x_ws[1] ]
def append_dict(wb_x,ws_x,data,mode='w',titels=[],search_col='',search_x='',creat_formated_table=True):#wb_path,sheet_name
    '''
    update:020316
    inputs
    ----------
        wb_x : str /obj(workbook)
            str= path fo file.
            if wb_x start by # make a auto_temp_file
        ws_x: str /obj(worksheet)
            str= sheet_name 
            .
        data : dict / list of dict
            dict= for put 1 row to table.
            list of dict = for put n row to table.
        mode: str=
            'a': append to end 
            'w': write = detete old) 
            's' :smart=search and replace 1 row
                in this case 'search_col' , 'serach_x' is used
                in other case 'search_col' , 'serach_x' not important and khali mimanand
        titels:list of str
            if wb_x =str and file is new 
                titels can be =[] ( not input) then pro : 
                    if type(data)=dict => get it from data.keys() 
                    if type(data)=list =>creat a serial number str strat from 0 
        search_col:str
            title of column for shearch in it
        search_x:str
            search text
        creat_formated_table:boolian
            آیا برنامه در فایل اکسل اطلاعات وارد شده را تبدیل به یک جدول فرمت شده اکسل تبدیل کند یا خیر
    outputs
    -------
        if mode=
            'a': append to end 
            'w': write = detete old) 
            's' :smart=search and replace 1 row
                in this case 'search_col' , 'serach_x' is used
                    replace row that row[search_col]=search_x
                in other case 'search_col' , 'serach_x' not important and can be skip (khali mimanand)
                            
                
        if wb & ws =str :
            if file not exist creat it else open it
                ....
            convert data to table 
            save & close
    '''
    # creat auto temp file name ساخت  نام فایل موقت اتوماتیک با توجه به ورودی در صورت نیاز
    if wb_x[0]=='#':
        import k_file
        wb_x=k_file.temp_file_name(name=wb_x[1:],ext='.xlsx')
    import tk_ui as ui
    def report_err(m1,x,cols_titles):
        ''' گزارش خطا'''     
        ui.msg (f"kxl.func(append_dict) \n error:in {m1} \n({x}) is not in columns\ncolumns={str(cols_titles)}\n press 1 key --> program wil be exit")
        #quit()
    def data_for_append_from_dic(titels,x_dic):   
        '''
        پیدا کردن عنوانهایی که در لیست عنوان موجود نیستند و اضافه کردن آنها به لیست عنوان و نیز اضافه کردن آنها به سطر اول فایل اکسل
        find all titel's from <x_dic> and append them to <titels>:list 
        '''        
        d_titels=[x for x in x_dic]
        for x in d_titels:
            if x not in titels:
                #report_err('check data',x,titels)
                ws.cell(row=1, column=len(titels)+1, value=x)
                titels.append(x)
        '''
        ساخت لیست مقادیر  جهت خروج
        '''   
        res=[]
        for x in titels:
            v=x_dic.get(x,'')
            if type(v) in [list,dict,tuple,set]: v=str(v)
            res.append(v)
        #res= [(x_dic.get(x,'')) for x in titels]       
        return res
    #-------------------------
    #data_titels:    
    d_titels=list(data.keys()) if type(data)==dict else list(data[0].keys()) 
    #
    if mode=='w': #delete all row
        if wb_x[0]!='#':
            import k_file
            k_file.file_delete_rcl(wb_x)    
    #
    if type(wb_x)==str:
        (wb,ws,new_file)=open_sheet(wb_x,ws_x)
        if new_file:
            if not titels:titels=d_titels
            ws.append(titels)
        else: #cols_titles
            titels=get_col_names(ws)
    else:    
        (wb,ws)=(wb_x,ws_x)
    #
    
    if mode in ['a','w']:#append
        if type(data)==dict: #1 row
            ws.append(data_for_append_from_dic(titels,data) )
        else:  #type(data)==list of dict => n row        
            for x_dic in data:
                ws.append(data_for_append_from_dic(titels,x_dic))
    elif mode=='s':#smart -serch and replce
        # این بخش از برنامه نیاز به بازبینی دارد
        data1={x.lower():data[x] for x in data}
        #ui.msg (f"kxl.append_dict \ndata1={data1}")
        
        #check all input_dict_keys is in table_titles
        for x in data1:
            if x not in cols_titles:
                report_err('check data',x,cols_titles)
        
        #creat sorted input_dict according table_titles 
        data2=[data1.get(x,'') for x in cols_titles]
        
        if search_col!='':
            if search_col in cols_titles:
                n_col=cols_titles.index(search_col)
                #ui.msg (f"kxl.append_dict \n ok:find \n({x}) is in columns\ncolumns={str(cols_titles)}\n position={n_col}")
                c,r=search_in_col_idx(ws,data1[search_col],n_col)
                if r==None: #search not found
                    ws.append(data2) 
                else:
                    print(f"kxl.append_dict-find data in {wb_x} in row:{r}" )
            else:
                report_err('search_col',search_col,cols_titles)
        else:
            ws.append(data2)
    #creat sorted input_dict according table_titles 
    ##data3=[str(data1.get(x,'')) for x in titels]        
    #ui.msg (f"kxl.append_dict \n ok:\ndta={data1}\ncolumns={str(cols_titles)}\n out={data3}")
    if type(wb_x)==str:
        if creat_formated_table:
            # table
            from openpyxl.worksheet.table import Table, TableStyleInfo
            tab = Table(displayName="Table1", ref="A1:"+col_addres(len(titels))+str(ws.max_row))
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            '''
            Table must be added using ws.add_table() method to avoid duplicate names.
            Using this method ensures table name is unque through out defined names and all other table name. 
            '''
            ws.add_table(tab)

        save_wb(wb,wb_x)
        import k_file
        k_file.launch_file(wb_x)
    else:    
        return wb,ws
def append_dict_test():
    xt='test2'
    if 1==1: #xt=='test1':
        data=[{'a':'abc','b':'123'},{'a':'qa2','b':'125'},
              {'a':'q2','b':'345'}]  
        append_dict(wb_x='c:\\temp\\test-py-out1.xlsx',ws_x='a',data=data)
    if xt=='test2':
        data={'a':'abc','b':'123'}
        append_dict(wb_x='#test1',ws_x='a',data=data)    
def append_list(wb_path,sheet_name,row='',rows=''):
    '''
    inputs
    ----------
        wb_path : str
            path o file.
        sheet_name : str
            .
        row : list of str (optional)
            for put 1 row to table
            also it is good for titel list
            if it=''(default) do nothing.
        rows : list of list (optional)
            for put multi row to table
            if it=''(default) do nothing.
    outputs
    -------
        if file not exist creat it else open it
        append row,rows sheet
        save & close
    Exam
    ------ 
    kxl.append_list('c:\\temp\\1.xlsx','0',['a','b','c'],[['1','2','3'],['4','5','6']])
    '''
    (wb,ws,file_is_new)=open_sheet(wb_path,sheet_name)
    if type(row)==list:
        ws.append(row)
    if type(rows)==list:    
        for row in rows:
            ws.append(row)
    save_wb(wb,wb_path)
def cols_list(ws,cols,col_name,search_v):
    n=cols.index(col_name)
    print(n)
    c=ws['a']
    c1=[x.value for x in c] 
    print (c1)
    return (search_v in c1)
def search_in_column(ws, search_string, column="A"):
    for row in range(1, ws.max_row + 1):
        coordinate = "{}{}".format(column, row)
        if ws[coordinate].value == search_string:
            return column, row
    return column, None
def search_in_col_idx(ws, search_string, col_idx=1):
    for row in range(1, ws.max_row + 1):
        if ws[row][col_idx].value == search_string:
            return col_idx, row
    return col_idx , None
def search_in_row_index(ws, search_string, row=1):
    for cell in ws[row]:
        if cell.value == search_string:
            return cell.column, row
    return None, row
def read(wb_path,sheet_name):
    """ read all data of sheet 01/09/14
        input:
        -------
            wb_path:str
                path of file for read
            sheet_name:str
                name of sheet for read
        output:
        -------
            titles:list of str
                list of table_titles
            rows:list of list
                Table_data in list format
            dic_list:list of dict
                table_data in dict format
            
    """
    wb,ws,new_file=open_sheet(wb_path,sheet_name)
    #
    def read_row(row):
        return [x.value for x in row]
    #---------------------------------
    titles=[]
    rows=[]
    dic_list=[]
    for row in ws:
        if not titles:
            titles=read_row(row)
        else:
            row_new=read_row(row)
            rows.append(row_new)
            dic_list.append({t:row_new[i] for i,t in enumerate(titles)})
    return titles,rows,dic_list
def save(data,file_name=r'c:\test\test.xlsx',titels=[]):
    ''' 01/09/14
        input:
        -------
            data:list of str / list fo dict
                
        file_name:str
    '''
    #print (data)
    wb,ws=open_sheet(file_name,sheet_name='0')
    #from openpyxl import load_workbook
    #wb = load_workbook(r'c:\test\test.xlsx')
    #ws = wb.active
    ws.append(["---"])
    if type(data[0])==dict:
        if not titels:titles=data[0].keys()
        for i,l in enumerate(titles):
            ws.cell(row=1, column=1+i).value=str(l)
        for i_r,row in enumerate(data):
            for i_c,col in enumerate(titles):
                ws.cell(row=2+i_r, column=1+i_c).value=str(row[col])    
    else:
        for i,l in enumerate(data):
            ws.cell(row=i+1, column=1).value=str(l)
    #ws.append(data)
    save_wb(wb,file_name)            
    
### not used
def _get_col_inf(x_ws,x_cn):
    '''
    get name and number of each colomn
    '''
    for i,x in enumerate(x_ws[1]): #range(100):
        c = str(x.value).lower()  #.cell(row=1 , column=i)
        if c in x_cn:
            x_cn[c]=i
            # print ("----ok --- {} is in {} ".format(c,i))
        else:    
            # print (c)
            pass
    for x in x_cn:
        if x_cn[x]<0:
            print('kxl.get_col_inf \n error  #### ({}) colomn not find in table'.format(x)) #,x_ws['name']) )       
def _find_in_col(x_ws,n_col,x_find):
    '''
    Parameters
    ----------
    x_ws : str
        worksheet object.
    n_col : number
        column number of x_ws for search whit in.
    x_find : str
        the target string for search.

    Returns
    -------
    Number
        the row number of found target
        if not found result=-1
        rows start from 0.

    '''
    r=0
    for row in x_ws.rows:
        if row[n_col].value == x_find:
            return r
        r+=1
    return -1  
def col_addres(n):
    '''
        return excel column name from col index exam( 1 =>a ,4=>d)
    '''    
    return chr(n+64) 
#####
#append_dict_test()
#print("ok")
#print (col_addres(1))
